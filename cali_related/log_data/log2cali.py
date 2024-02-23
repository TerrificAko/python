import re
import openpyxl
import os

path = r"C:\proj\python\data"
os.chdir(path)


tia_s = re.compile("THIS CASE J:\s*(\d+) M:\s*(\d+)" )
tia_d = re.compile(".*tia case :\d+,absolute value dc_i : (-*\d+) , value dc_q : (-*\d+)" )
vga_d = re.compile(".*vga case:\d+ ,absolute value dc_i_tmp : (-*\d+) , value dc_q_tmp : (-*\d+) ,")

def reg_match3(line, data):
    c = re.compile(r"reg(\w+)=([^ . , \n]+)")
    matches = c.findall(line)
    for match in matches:
        data.append((match[0], match[1]))
    return bool(matches)


pre_reg = ['0x1180',"0x11bc","0x11f8","0x1234"]

arr_tia = []
arr_vga = []
arr_cali_pre = []
with open("./test1.DAT",'r') as f:
    isTia = False
    isVga = False
    iscali_pre = False
    tia_index = 0

    data_tia_dc = []
    data_tia_post = []
    data_vga_dc = []
    data_vga_post = []
    data_cali_pre = []

    lines = f.readlines()
    for index,line in enumerate(lines):
        tia_match_s = tia_s.match(line)
        if tia_match_s :
            tia_index = (int(tia_match_s.group(1)) +1) * (int(tia_match_s.group(2)) +1)
            # print(tia_match_s.group(1),tia_match_s.group(2),tia_index,index)
            iscali_pre = False
            isTia = False
            isVga = False
            if(data_cali_pre):
                arr_cali_pre.append(data_cali_pre)
            data_cali_pre = []
            # print(data_cali_pre)
            print(data_tia_dc)
            print(data_tia_post)

            if data_tia_dc:
                arr_tia.append([data_tia_dc,data_tia_post])
            data_tia_dc = []
            data_tia_pre = []
            data_tia_post = []
            # print(data_vga_dc)
            # print(data_vga_post)
            if data_vga_dc:
                arr_vga.append([data_vga_dc,data_vga_post])
            data_vga_dc = []
            data_vga_post = []
            # if tia_index > 10:
            #     break
            continue

        if "cali pre" in line:
            iscali_pre =  True
            isTia = False
            isVga = False

        if "rx0 tia cali begin" in line:
            iscali_pre =  False
            isTia = True
            isVga = False

        if "rx0 vga calibration begin" in line:
            iscali_pre =  False
            isTia = False
            isVga = True

        
        if iscali_pre:
            reg_match3(line, data_cali_pre)
            # for i in pre_reg:
            #     reg_match(i,line,data_cali_pre)

        if isTia:
            mt = tia_d.match(line)
            if mt:
                data_tia_dc.append((mt.group(1),mt.group(2)))
                # reg_match2(lines[index+1],data_tia_post)
                reg_match3(lines[index+1],data_tia_post)

        if isVga:
            mt = vga_d.match(line)
            if mt:
                data_vga_dc.append((mt.group(1),mt.group(2)))
                reg_match3(line,data_vga_post)
    # print(data_cali_pre)
    if (data_cali_pre):
        arr_cali_pre.append(data_cali_pre)
    data_cali_pre = []
    # print(data_cali_pre)
    # print(arr_cali_pre)
    # print(data_tia_post)


    if data_tia_dc:
        arr_tia.append([data_tia_dc, data_tia_post])
    data_tia_dc = []
    data_tia_pre = []
    data_tia_post = []
    # print(data_vga_dc)
    # print(data_vga_post)
    if data_vga_dc:
        arr_vga.append([data_vga_dc, data_vga_post])
    data_vga_dc = []
    data_vga_post = []

# print(index)
# print(arr_cali_pre)
# print(arr_tia)
# print(arr_vga)

# print(arr_tia[0][1][2])
# print(arr_tia[0][1][2][1])
        
# 打开Excel文件
workbook = openpyxl.load_workbook('output.xlsx')
# 选择第一个工作表
worksheet = workbook.active
for i in range(tia_index):
    #col below
    #tia
    for j in range(len(arr_tia[i])):
        # dc & regvalue
        for k in range(len(arr_tia[i][j])):
            #which tuple
            for m in range(len(arr_tia[i][j][k])):
                #Element within tuple
                if(j == 0):
                    #dc
                    row_index = 2 + k*2 + m
                    # print(row_index)
                    worksheet.cell(row=row_index, column=i+5).value = arr_tia[i][j][k][m]
                    # print(arr_tia[i][j][k][m])

                if(j == 1 & m == 1):
                    #regvalue
                    row_index = 54 + k
                    worksheet.cell(row=row_index, column=i + 5).value = arr_tia[i][j][k][m]

    # vga
    for j in range(len(arr_cali_pre[i])):
        #arr_cali_pre which tuple
        # print(arr_cali_pre[i][j])
        for k in range(len(arr_cali_pre[i][j])):
            # arr_cali_pre Element within tuple
            # print(arr_cali_pre[i][j][k] , i,j,k)
            if (j < 2 ):
                if (k == 1):
                    row_index = 136 + j
                    worksheet.cell(row=row_index, column=i + 5).value = arr_cali_pre[i][j][k]
                    # print(arr_cali_pre[i][j][k], i, j, k)

            if (j > 1 ):
                if (k == 1):
                    row_index = 166 + j - 2
                    worksheet.cell(row=row_index, column=i + 5).value = arr_cali_pre[i][j][k]
                    # print(arr_cali_pre[i][j][k], i, j, k)

    for j in range(len(arr_vga[i])):
        # dc & regvalue
        # print(arr_vga[i][j])
        for k in range(len(arr_vga[i][j])):
            #which tuple
            # print(arr_vga[i][j][k])
            for m in range(len(arr_vga[i][j][k])):
                #Element within tuple
                if(j == 0):
                    #dc
                    row_index = 80 + k * 2 + m
                    # print(row_index)
                    worksheet.cell(row=row_index, column=i+5).value = arr_vga[i][j][k][m]
                    # print(arr_vga[i][j][k][m])
                if(j == 1):
                    #reg value
                    if(k < 28 ):
                        if(m == 1):
                            row_index = 138 + k
                            # print(row_index)
                            worksheet.cell(row=row_index, column=i + 5).value = arr_vga[i][j][k][m]
                            # print(arr_vga[i][j][k][m])
                    if (k > 27):
                        if(m == 1):
                            row_index = 168 + k - 28
                            # print(row_index)
                            worksheet.cell(row=row_index, column=i + 5).value = arr_vga[i][j][k][m]
                            # print(arr_vga[i][j][k][m])


# 保存Excel文件
workbook.save('output.xlsx')







        
