import re
import openpyxl
import os

class LogParser:
    def __init__(self, path, endname):

        self.path = path
        self.endname = endname
        self.file_dict = {}
        self.cali_st = None
        self.isCalipre = False
        self.isTia = False
        self.isVga = False
        self.wb = openpyxl.Workbook()

    def find_files(self):
        files = []
        for f in os.listdir(self.path):
            if f.endswith(self.endname):
                files.append(f)
        return files

    def load_all_data(self):
        for file in self.find_files():
            self.file_dict.setdefault(file, {})

            file_path = os.path.join(self.path, file)
            # print(file_path)
            # print(self.file_dict)
            self.extract_data(file_name=file, file_path_name=file_path)

    def extract_data(self, file_name, file_path_name):
        # get_file_content by func
        file_content = self.load_conent(file_path_name)
        self.match_cfg()
        self.data_cali_pre = []
        self.data_tia_reg = []
        self.data_vga_reg = []
        adc_case_fix = ""
        for index, line in enumerate(file_content):
            cali_st_flag = self.cali_st.match(line)
            if cali_st_flag:
                adc_case_fix = "M: {}, J: {}".format(cali_st_flag.group(1), cali_st_flag.group(2))
                self.file_dict.setdefault(file_name, {}).setdefault(adc_case_fix, {})
                self.isCalipre = False
                self.isTia = False
                self.isVga = False

                self.data_cali_pre = []
                self.data_tia_reg = []
                self.data_vga_reg = []

            if "adc calibration begin" in line:
                self.isCalipre = True
                self.isTia = False
                self.isVga = False

            if "rx0 tia cali begin" in line:
                self.isCalipre = False
                self.isTia = True
                self.isVga = False

            if "rx0 vga calibration begin" in line:
                self.isCalipre = False
                self.isTia = False
                self.isVga = True

            if self.isCalipre:
                self.reg_match(line, self.data_cali_pre)
                # print(self.data_cali_pre)
                self.file_dict.setdefault(file_name, {}).setdefault(adc_case_fix, {}).setdefault('cali_pre',self.data_cali_pre)
                # self.file_dict[file_name][adc_case_fix]["cali_pre"] = self.data_cali_pre
            if self.isTia:
                Tia_data_flag = self.tia_d.match(line)
                if Tia_data_flag:
                    self.data_tia_reg = []
                    TIA_RX_fix = "TIA_RX{}".format(Tia_data_flag.group(1))
                    tia_case_fix = "tia case{}".format(Tia_data_flag.group(2))
                    self.reg_match(file_content[index + 1], self.data_tia_reg)
                    tia_case_data = [Tia_data_flag.group(3), Tia_data_flag.group(4), self.data_tia_reg]
                    # print(tia_case_data)
                    self.file_dict.setdefault(file_name, {}).setdefault(adc_case_fix, {}).setdefault('tia_data',{}).setdefault(TIA_RX_fix, {}).setdefault(tia_case_fix, tia_case_data)
                    # self.file_dict[file_name][adc_case_fix][TIA_RX_fix][tia_case_fix] = tia_case_data
            if self.isVga:
                Vga_data_flag = self.vga_d.match(line)
                if Vga_data_flag:
                    self.data_vga_reg = []
                    VGA_RX_fix = "VGA_RX{}".format(Vga_data_flag.group(1))
                    vga_case_fix = "vga case{}".format(Vga_data_flag.group(2))
                    self.reg_match(file_content[index], self.data_vga_reg)
                    vga_case_data = [Vga_data_flag.group(3), Vga_data_flag.group(4), self.data_vga_reg]
                    self.file_dict.setdefault(file_name, {}).setdefault(adc_case_fix, {}).setdefault('vga_data',{}).setdefault(VGA_RX_fix, {}).setdefault(vga_case_fix, vga_case_data)

        # print(self.file_dict)

    def reg_match(self, line, data):
        c = re.compile(r"reg(\w+)=([^ . , \n]+)")
        matches = c.findall(line)
        for match in matches:
            data.append((match[0], match[1]))
        return bool(matches)

    def match_cfg(self):
        self.cali_st = re.compile("THIS CASE J:\s*(\d+) M:\s*(\d+)")
        self.tia_d = re.compile("RX(\d+) tia case :(\d+),absolute value dc_i : (-*\d+) , value dc_q : (-*\d+)")
        self.vga_d = re.compile("RX(\d+) vga case:(\d+) ,absolute value dc_i_tmp : (-*\d+) , value dc_q_tmp : (-*\d+) ,")

    def load_conent(self, file_path):
        content = []
        with open(file_path, 'r') as f:
            content = f.readlines()
        return content

    def save_to_excel(self, file_name):
        self.output_file_name = os.path.join(self.path, file_name)
        print(self.output_file_name)
        # self.workbook = openpyxl.Workbook(self.output_file_name) # 创建新的工作簿
        # self.workbook.save(self.output_file_name)  # 保存工作簿到文件
        self.wb.save(self.output_file_name)
        # self.workbook = openpyxl.load_workbook(self.output_file_name)
        for file, file_data in self.file_dict.items():
            sheet_name = file
            # print(sheet_name)
            self.ws = self.wb.create_sheet(title=sheet_name)
            adc_case_name = []
            self.index_col = 4
            for adc_case, adc_case_data in file_data.items():
                adc_case_name.append(adc_case)
            # print(adc_case_name)
                self.index_col = self.index_col + 1
                for once_key, once_cali_data in adc_case_data.items():
                    # print(once_key)
                    if once_key == "cali_pre":
                        self.write_cali_pre(once_cali_data)

                    if once_key == "tia_data":
                        self.write_tia_data(once_cali_data)

                    if once_key == "vga_data":
                        self.write_vga_data(once_cali_data)

            self.write_column(adc_case_name)
            self.wb.save(self.output_file_name)

    def write_column(self, list_name):
        for index, name in enumerate(list_name):
            # print(index)
            self.ws.cell(row=1, column=index+5).value = name
            # self.worksheet.cell(row=1, column=index + 5, value=name)

        # self.workbook.save(self.output_file_name)

    def write_cali_pre(self, once_cali_data):
        for index, cali_data in enumerate(once_cali_data):
            # print(index)
            if index < 2:
                self.row_index = index + 136
                # print(cali_data[1])
                self.ws.cell(row=self.row_index, column=self.index_col).value = cali_data[1]
            if index > 1:
                self.row_index = index + 166 - 2
                self.ws.cell(row=self.row_index, column=self.index_col).value = cali_data[1]


    def write_tia_data(self, once_tia_data):
        for tia_rx, tia_rx_data in once_tia_data.items():
            if tia_rx == "TIA_RX0":
                self.row_index_dc = 2
                self.row_index_reg = 54

            if tia_rx == "TIA_RX1":
                self.row_index_dc = 28
                self.row_index_reg = 67
            # print(tia_rx)
            for index, (tia_case_index, tia_case_data) in enumerate(tia_rx_data.items()):
                # print(index)
                # print(tia_case_data)
                # print(tia_case_data[2][0][1])
                self.row_index = index * 2 + self.row_index_dc
                self.ws.cell(row=self.row_index, column=self.index_col).value = tia_case_data[0]
                self.ws.cell(row=(self.row_index + 1), column=self.index_col).value = tia_case_data[1]
                self.row_index = index + self.row_index_reg
                self.ws.cell(row=self.row_index, column=self.index_col).value = tia_case_data[2][0][1]

    def write_vga_data(self, once_vga_data):
        for vga_rx, vga_rx_data in once_vga_data.items():
            if vga_rx == "VGA_RX0":
                self.row_index_dc = 80
                self.row_index_reg = 138

            if vga_rx == "VGA_RX1":
                self.row_index_dc = 108
                self.row_index_reg = 168

            for index, (vga_case_index, vga_case_data) in enumerate(vga_rx_data.items()):
                # print(index)
                # print(vga_case_data)
                # print(vga_case_data[2][0][1])
                # print(vga_case_data[2][1][1])
                self.row_index = index * 2 + self.row_index_dc
                self.ws.cell(row=self.row_index, column=self.index_col).value = vga_case_data[0]
                self.ws.cell(row=(self.row_index + 1), column=self.index_col).value = vga_case_data[1]

                self.row_index = index * 2 + self.row_index_reg
                self.ws.cell(row=self.row_index, column=self.index_col).value = vga_case_data[2][0][1]
                self.ws.cell(row=(self.row_index + 1), column=self.index_col).value = vga_case_data[2][1][1]


def main():
    parser = LogParser(path='C:\proj\python\data\ic_data', endname='.DAT')
    parser.load_all_data()
    parser.save_to_excel("output.xlsx")


if __name__ == '__main__':
    main()
