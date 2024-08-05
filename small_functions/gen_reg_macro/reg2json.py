import openpyxl
import json
# wb = openpyxl.load_workbook(
#     './json2reg/ZN2014_Register_Mapping_aon.xlsx', data_only=True)

wb = openpyxl.load_workbook(
    # './json2reg/ana2soc_0104.xlsx', data_only=True)
    # './json2reg/ana2soc_3107_0308.xlsx', data_only=True)
    './json2reg/ZN3107_ana2soc_aon_analog_03212024.xlsx', data_only=True)


# wb = openpyxl.load_workbook(
#     './json2reg/ZN2014_Register_Mapping.xlsx', data_only=True)
reset_test = 12
reset = 9
description = 8
name = 7
accType = 6
bitwidth = 4
endbit = 3
startbit = 2
offset1 = 1
offset0 = 0

data = []
flag = 0
fields = []

# add test mode for ana
TEST_MODE=True

reserved = {
    "accType": "NA",
    "name": "--",
    "width": 1,
    "reset": 0,
    "reset_test": -1,
    "doc": "Reserved"
}
# MAX_BITSWIDTH = 8
MAX_BITSWIDTH = 32
endsbits = MAX_BITSWIDTH -1 
for sheetname in wb.sheetnames:
    # if sheetname.startswith('reg'):
    # if sheetname.startswith('AON_CFG'):
    if sheetname.startswith('ana2soc'):
        sheet = wb[sheetname]
        # prefix = sheetname[3:]
        cnt = 2
        for row in sheet.iter_rows(min_row=2):
            # print(cnt)
            cnt += 1
            if row[offset1].value != None:
                # if row[offset0].value != prefix:
                #     print(row)
                #     print("offset not equal")
                if flag != 0:
                    if endsbits != -1:
                        reserved['width'] = endsbits + 1
                        fields.append(reserved.copy())
                    reg['fields'] = fields.copy()[::-1]
                    # print(reg['name'])
                    data.append(reg.copy())
                    fields = []
                endsbits = MAX_BITSWIDTH -1
                reg = {}
                # print(row[offset0].value, row[offset1].value)
                reg['addr'] = int(row[offset0].value+row[offset1].value, 16)
                reg['name'] = row[name].value
                reg['doc'] = row[description].value
                flag = 1
            else:
                if row[endbit].value != endsbits:
                    print(cnt,"======",reg['name'],"====",row[endbit].value)
                    # print(dir(row[endbit].value))
                    reserved['width'] = endsbits - row[endbit].value
                    fields.append(reserved.copy())
                    endsbits = row[endbit].value
                if row[startbit].value > row[endbit].value:
                    print("bitwidth not equal")
                    print(cnt,reg['name'],row)
                    exit
                if endsbits < row[endbit].value:
                    print("Unexpect bit order.")
                    print(cnt,reg['name'],row)
                    exit
                reset_value = row[reset].value
                if reset_value == None:
                    reset_value = 0

                reset_test_value = row[reset_test].value
                if reset_test_value == None:
                    reset_test_value = -1
                if row[name].value.lower() == "reserved":
                    reserved['width'] = row[endbit].value-row[startbit].value+1
                    fields.append(reserved.copy())
                else:
                    fields.append({
                        'accType': row[accType].value,
                        'name': row[name].value,
                        'width': row[endbit].value-row[startbit].value+1,
                        'reset': reset_value,
                        'reset_test': reset_test_value,
                        'doc': row[description].value})
                endsbits -= row[endbit].value - row[startbit].value + 1

if endsbits > 0:
    reserved['width'] = endsbits
    fields.append(reserved.copy())
reg['fields'] = fields.copy()[::-1]
data.append(reg.copy())
# exit()
# with open("aon_reg.json", "w") as f:
# with open("uwb_reg.json", "w") as f:
with open("ana2soc_3107_0322.json", "w") as f:
    json.dump(data, f)
